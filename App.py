from flask import Flask, request, send_file, jsonify, session, render_template_string, redirect, url_for
from functools import wraps
import pdfplumber
import re
import os
import tempfile
import smtplib
import requests
import json
import traceback
import urllib3
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys

# ── Suprime avisos de SSL (igual ao script 2) ──────────────────────────────
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

app = Flask(__name__, static_folder='.', static_url_path='')

# ════════════════════════════════════════════════════════════════════════════
# ─── CONFIGURAÇÃO / CREDENCIAIS (via variáveis de ambiente) ─────────────────
# ════════════════════════════════════════════════════════════════════════════
# IMPORTANTE: defina estas variáveis de ambiente antes de rodar o servidor.
# Nunca deixe senhas reais hardcoded no código-fonte.
#
#   export FLASK_SECRET_KEY="algo-aleatorio-e-longo"
#   export SENHA_ADMIN="sua-senha"
#   export EMAIL_ADMIN="seu-email@gmail.com"
#   export EMAIL_SENDER="seu-email@gmail.com"
#   export EMAIL_PASSWORD="sua-senha-de-app-do-gmail"
#   export SAP_SL_URL="https://b1.ativy.com:51032/b1s/v1"   ← mesmo do script 2
#   export SAP_COMPANY_DB="NOME_DA_BASE"
#   export SAP_USER="usuario_sap"
#   export SAP_PASSWORD="senha_sap"

app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'sua_chave_secreta_aqui_mude_em_producao')

lista = {}

SENHA_ADMIN  = os.environ.get('SENHA_ADMIN', '2604')
EMAIL_ADMIN  = os.environ.get('EMAIL_ADMIN', 'gregorydeabreu89@gmail.com')
ARQUIVO_TENTATIVAS = 'tentativas_login.json'

SENHAS_HONEYPOT = [
    "Conseguimos", "admin", "123456", "password", "root",
    "senha123", "12345678", "admin123", "qwerty", "123123", "passw0rd",
]

NOMES_ESPECIAIS_CLT = ['FABIANO LUIS MANFRON MORO', 'JOAO GUILHERME SILVEIRA', 'KAUA DOS SANTOS VIANA']

# Nome do funcionário que recebe a regra especial de PRO-LABORE no lançamento contábil.
NOME_PROLABORE_FABIANO = 'FABIANO LUIS MANFRON MORO'

SMTP_SERVER   = "smtp.gmail.com"
SMTP_PORT     = 587
EMAIL_SENDER  = os.environ.get('EMAIL_SENDER', 'gregorydeabreu89@gmail.com')
EMAIL_PASSWORD = os.environ.get('EMAIL_PASSWORD', '')  # defina via variável de ambiente

# ── Constantes do lançamento contábil ───────────────────────────────────────
DATA_LANC_PADRAO = datetime(2025, 5, 31)


def resolver_datas_sap(data_lanc_str=None):
    """
    Recebe a data de lançamento escolhida no site (string 'YYYY-MM-DD') e calcula:
      - data_lanc   : data de lançamento (escolhida pelo usuário)
      - data_venc   : data de vencimento = data de lançamento + 5 dias
      - data_doc    : data do documento  = igual à data de lançamento
      - competencia : mês seguinte ao da data de lançamento, formato 'MM.AAAA'

    Se nenhuma data for enviada, usa a data padrão (fallback).
    """
    if data_lanc_str:
        try:
            data_lanc = datetime.strptime(data_lanc_str, '%Y-%m-%d')
        except (ValueError, TypeError):
            data_lanc = DATA_LANC_PADRAO
    else:
        data_lanc = DATA_LANC_PADRAO

    data_venc = data_lanc + timedelta(days=5)
    data_doc  = data_lanc

    mes = data_lanc.month + 1
    ano = data_lanc.year
    if mes > 12:
        mes = 1
        ano += 1
    competencia = f'{mes:02d}.{ano}'

    return data_lanc, data_venc, data_doc, competencia


# ── Configuração do SAP Business One Service Layer ─────────────────────────
# Nomes das variáveis de ambiente IDÊNTICOS ao script 2 (PJ):
#   CompanyDB   → nome do banco SAP  (mesmo do script 2)
#   SAP_USER    → usuário SAP
#   SAP_PASSWORD→ senha SAP
#   SAP_SL_URL  → URL base (padrão já aponta para o servidor Ativy)
_sap_url_raw = os.environ.get("SAP_SL_URL", "https://b1.ativy.com:51032/b1s/v1").rstrip("/")
for _ep in ("/JournalEntries", "/Login", "/JournalVouchers"):
    if _sap_url_raw.endswith(_ep):
        _sap_url_raw = _sap_url_raw[:-len(_ep)]
        break
SAP_SL_URL = _sap_url_raw.rstrip("/")
SAP_COMPANY_DB = os.environ.get('CompanyDB', '')    # ← igual ao script 2: os.getenv("CompanyDB")
SAP_USER       = os.environ.get('SAP_USER', '')     # ← igual ao script 2: os.getenv("SAP_USER")
SAP_PASSWORD   = os.environ.get('SAP_PASSWORD', '') # ← igual ao script 2: os.getenv("SAP_PASSWORD")
# verify=False igual ao script 2 (suprime verificação de certificado autoassinado)
SAP_VERIFY_SSL = False

# Código da filial (Branch / BPLID) usada nos lançamentos contábeis.
# O SAP exige esse campo em TODAS as linhas quando a base tem filiais habilitadas.
# Ajuste via variável de ambiente SAP_BPLID se a filial padrão for diferente de 1.
BPL_ID_PADRAO = 3


# ════════════════════════════════════════════════════════════════════════════
# ─── LOGIN / SEGURANÇA ───────────────────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

def carregar_tentativas():
    if os.path.exists(ARQUIVO_TENTATIVAS):
        try:
            with open(ARQUIVO_TENTATIVAS, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {}
    return {}

def salvar_tentativas(dados):
    with open(ARQUIVO_TENTATIVAS, 'w', encoding='utf-8') as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)

def obter_ip_cliente():
    if request.headers.getlist("X-Forwarded-For"):
        return request.headers.getlist("X-Forwarded-For")[0]
    return request.remote_addr

def obter_info_ip(ip):
    if ip.startswith('127.') or ip.startswith('192.168.') or ip.startswith('10.'):
        return {"ip": ip, "ip_publico": ip, "pais": "Acesso Local",
                "estado": "Rede Privada", "cidade": "Localhost/LAN", "isp": "Local"}
    try:
        resposta = requests.get(f"https://ipapi.co/{ip}/json/", timeout=5)
        if resposta.status_code == 200:
            dados = resposta.json()
            return {"ip": ip, "pais": dados.get("country_name", "Desconhecido"),
                    "estado": dados.get("region", "Desconhecido"),
                    "cidade": dados.get("city", "Desconhecido"),
                    "isp": dados.get("org", "Desconhecido"), "ip_publico": ip}
    except Exception as e:
        print(f"⚠️  Erro IP: {e}", flush=True)
    return {"ip": ip, "pais": "Não disponível", "estado": "Não disponível",
            "cidade": "Não disponível", "isp": "Não disponível", "ip_publico": ip}

def enviar_email_alerta(ip, info_ip, email_destino):
    try:
        assunto = "⚠️  ALERTA DE SEGURANÇA: Tentativa de Login Falhada"
        corpo_html = f"""<html><body style="font-family:Arial">
        <h2 style="color:#d32f2f">⚠️ ALERTA DE SEGURANÇA</h2>
        <p><b>IP:</b> {info_ip['ip_publico']}</p>
        <p><b>País:</b> {info_ip['pais']}</p>
        <p><b>Estado:</b> {info_ip['estado']}</p>
        <p><b>Cidade:</b> {info_ip['cidade']}</p>
        <p><b>ISP:</b> {info_ip['isp']}</p>
        <p><b>Data/Hora:</b> {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')}</p>
        </body></html>"""
        mensagem = MIMEMultipart("alternative")
        mensagem["Subject"] = assunto
        mensagem["From"] = EMAIL_SENDER
        mensagem["To"] = email_destino
        mensagem.attach(MIMEText(corpo_html, "html"))
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as s:
            s.starttls(); s.login(EMAIL_SENDER, EMAIL_PASSWORD)
            s.sendmail(EMAIL_SENDER, email_destino, mensagem.as_string())
        return True
    except Exception as e:
        print(f"❌ Erro alerta: {e}", flush=True)
        return False

def enviar_alerta_honeypot(ip, senha_tentada, info_ip, email_destino):
    try:
        assunto = "🚨 ALERTA CRÍTICO: TENTATIVA DE INVASÃO!"
        corpo_html = f"""<html><body style="font-family:Arial">
        <h2 style="color:#c41c3b">🚨 INVASÃO DETECTADA</h2>
        <p><b>Senha armadilha usada:</b> {senha_tentada}</p>
        <p><b>IP:</b> {ip} | <b>Cidade:</b> {info_ip['cidade']} | <b>País:</b> {info_ip['pais']}</p>
        <p><b>Data/Hora:</b> {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')}</p>
        </body></html>"""
        mensagem = MIMEMultipart("alternative")
        mensagem["Subject"] = assunto
        mensagem["From"] = EMAIL_SENDER
        mensagem["To"] = email_destino
        mensagem.attach(MIMEText(corpo_html, "html"))
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as s:
            s.starttls(); s.login(EMAIL_SENDER, EMAIL_PASSWORD)
            s.sendmail(EMAIL_SENDER, email_destino, mensagem.as_string())
        return True
    except Exception as e:
        print(f"❌ Erro honeypot: {e}", flush=True)
        return False

HTML_FILE = os.path.join(os.path.dirname(__file__), 'index.html')

CODIGOS_CLT = {
    "BRUNO RAFAEL DOS SANTOS DE MORAES": "F001579",
    "DIEGO SAMPAIO DA SILVA": "F001586",
    "EDUARDO SILVA ANTUNES": "F001576",
    "FILIPI BOM SIMON DE SOUZA": "F001741",
    "HENRIQUE MARLEY MATEUS KRUTLI": "F001989",
    "HENRIQUE MARTINS DE FREITAS":"F001675",
    "EDUARDO SILVA ANTUNES": "F001576",
    "JAMES DE OLIVEIRA PINTO": "F001785",
    "JULIO CESAR PEREIRA FRANCISCO": "F001962",
    "KAUA DOS SANTOS VIANA": "F000009",
    "LUCAS CABRAL BOTT": "F001735",
    "LUCAS DO CANTO": "F001610",
    "MATEUS OSTROWSKI": "F001875",
    "MIQUEIAS DOS SANTOS DA GAMA": "F001590",
    "RENATA PACHECO BARBOZA": "F001871",
    "ROGER VIEIRA BIANCHIN": "F002134",
    "VICTOR NETO TORETTI": "F001588",
    "FABIANO LUIS MANFRON MORO": "F001592"

    # Adicione quantos funcionários desejar...
}

@app.route('/login', methods=['GET', 'POST'])
def login():
    print(f"\n🔐 /LOGIN {request.method}", flush=True)
    if request.method == 'POST':
        senha_fornecida = request.form.get('senha', '')
        ip_cliente = obter_ip_cliente()

        if senha_fornecida in SENHAS_HONEYPOT:
            tentativas = carregar_tentativas()
            tentativas[ip_cliente] = {'tentativas': 999, 'bloqueado': True,
                'primeira_tentativa': datetime.now().isoformat(),
                'ultima_tentativa': datetime.now().isoformat(),
                'honeypot': True, 'senha_usada': senha_fornecida}
            salvar_tentativas(tentativas)
            try:
                info_ip = obter_info_ip(ip_cliente)
                enviar_alerta_honeypot(ip_cliente, senha_fornecida, info_ip, EMAIL_ADMIN)
            except Exception as e:
                print(f"❌ {e}", flush=True)
            return render_template_string("""<!DOCTYPE html><html><head><meta charset="UTF-8">
                <title>Bloqueado</title><style>body{font-family:Arial;background:linear-gradient(135deg,#667eea,#764ba2);
                min-height:100vh;display:flex;align-items:center;justify-content:center;margin:0}
                .c{background:white;padding:40px;border-radius:10px;text-align:center;max-width:500px}
                h1{color:#d32f2f}.w{background:#fff3cd;border:2px solid red;padding:15px;border-radius:5px;
                margin:20px 0;color:#660000;font-weight:bold}</style></head>
                <body><div class="c"><h1>🚨 Acesso Bloqueado Permanentemente</h1>
                <div class="w">⚠️ TENTATIVA DE INVASÃO DETECTADA!<p>Seus dados foram registrados.</p>
                </div></div></body></html>""")

        tentativas = carregar_tentativas()
        if ip_cliente in tentativas and tentativas[ip_cliente]['bloqueado']:
            return render_template_string("""<!DOCTYPE html><html><head><meta charset="UTF-8">
                <title>Bloqueado</title><style>body{font-family:Arial;background:linear-gradient(135deg,#667eea,#764ba2);
                min-height:100vh;display:flex;align-items:center;justify-content:center;margin:0}
                .c{background:white;padding:40px;border-radius:10px;text-align:center;max-width:500px}
                h1{color:#d32f2f}.w{background:#fff3cd;border:1px solid #ffc107;padding:15px;border-radius:5px;
                margin:20px 0;color:#856404}</style></head>
                <body><div class="c"><h1>🔒 Acesso Bloqueado</h1>
                <div class="w">⚠️ Tentativa não autorizada detectada!
                <p>Alerta enviado ao administrador.</p></div>
                <div style="background:#e3f2fd;border:1px solid #2196f3;padding:15px;border-radius:5px;
                color:#0d47a1;font-size:14px">IP: """ + ip_cliente + """<br>
                Data/Hora: """ + datetime.now().strftime('%d/%m/%Y %H:%M:%S') + """</div>
                </div></body></html>""")

        if senha_fornecida == SENHA_ADMIN:
            session['logado'] = True
            session['ip_login'] = ip_cliente
            session['hora_login'] = datetime.now().isoformat()
            if ip_cliente in tentativas:
                del tentativas[ip_cliente]
                salvar_tentativas(tentativas)
            return redirect(url_for('index'))
        else:
            if ip_cliente not in tentativas:
                tentativas[ip_cliente] = {'tentativas': 0, 'bloqueado': False,
                    'primeira_tentativa': datetime.now().isoformat()}
            tentativas[ip_cliente]['tentativas'] += 1
            tentativas[ip_cliente]['ultima_tentativa'] = datetime.now().isoformat()
            tentativas[ip_cliente]['bloqueado'] = True
            salvar_tentativas(tentativas)
            try:
                info_ip = obter_info_ip(ip_cliente)
                enviar_email_alerta(ip_cliente, info_ip, EMAIL_ADMIN)
            except Exception as e:
                print(f"❌ {e}", flush=True)
            return render_template_string("""<!DOCTYPE html><html><head><meta charset="UTF-8">
                <title>Bloqueado</title><style>body{font-family:Arial;background:linear-gradient(135deg,#667eea,#764ba2);
                min-height:100vh;display:flex;align-items:center;justify-content:center;margin:0}
                .c{background:white;padding:40px;border-radius:10px;text-align:center;max-width:500px}
                h1{color:#d32f2f}.w{background:#fff3cd;border:1px solid #ffc107;padding:15px;border-radius:5px;
                margin:20px 0;color:#856404}</style></head>
                <body><div class="c"><h1>🔒 Acesso Bloqueado</h1>
                <div class="w">⚠️ Alerta de Segurança Enviado
                <p>Um email com seus dados foi enviado ao administrador.</p>
                </div></div></body></html>""")

    return render_template_string("""<!DOCTYPE html><html><head><meta charset="UTF-8">
        <meta name="viewport" content="width=device-width,initial-scale=1.0">
        <title>Login - Folha de Pagamento</title>
        <style>body{font-family:Arial;background:linear-gradient(135deg,#667eea,#764ba2);
        min-height:100vh;display:flex;align-items:center;justify-content:center;margin:0}
        .c{background:white;padding:40px;border-radius:10px;box-shadow:0 10px 25px rgba(0,0,0,.2);
        max-width:400px;width:100%}h1{color:#333;text-align:center;margin:0 0 10px}
        .sub{color:#999;text-align:center;font-size:14px;margin-bottom:30px}
        form{display:flex;flex-direction:column}label{color:#333;font-weight:bold;margin-bottom:5px}
        input{padding:12px;margin-bottom:20px;border:1px solid #ddd;border-radius:5px;font-size:14px}
        input:focus{outline:none;border-color:#667eea}
        button{background:linear-gradient(135deg,#667eea,#764ba2);color:white;padding:12px;
        border:none;border-radius:5px;font-size:16px;font-weight:bold;cursor:pointer}
        button:hover{opacity:.9}</style></head>
        <body><div class="c">
        <h1>🔐 Folha de Pagamento</h1><p class="sub">Acesso Restrito</p>
        <form method="POST">
        <label for="senha">Senha:</label>
        <input type="password" id="senha" name="senha" placeholder="Digite a senha" required autofocus>
        <button type="submit">Entrar</button></form>
        </div>
        </body></html>""")

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

def requer_autenticacao(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get('logado'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return wrapper

@app.route('/')
@requer_autenticacao
def index():
    return send_file(HTML_FILE)

def safe_float(v):
    try:
        return float(str(v).replace('.', '').replace(',', '.'))
    except:
        return 0.0


def valor_resumo(block, rotulo):
    m = re.search(rotulo + r':\s*([-\d.,]+)', block, re.IGNORECASE)
    return safe_float(m.group(1)) if m else 0.0


def valor_rubrica(block, codigo, descricao, natureza):
    padrao = rf'\b{codigo}\s+{descricao}\b.*?([\d.,]+)\s*{natureza}\b'
    m = re.search(padrao, block, re.IGNORECASE | re.DOTALL)
    return safe_float(m.group(1)) if m else 0.0


def aplicar_regra_fabiano(emp, block):
    if emp['nome'].strip().upper() != 'FABIANO LUIS MANFRON MORO':
        return

    m_pl = re.search(r'\b100\s+PRO-LABORE\b.*?(\d+:\d+)\s+([\d.,]+)\s*P', block, re.IGNORECASE)
    emp['pro_labore'] = safe_float(m_pl.group(2)) if m_pl else 0.0

    m_inss_emp = re.search(r'\b843\s+DESC\.INSS\s+EMPREGADOR\b.*?([\d.,]+)\s+([\d.,]+)\s*D', block, re.IGNORECASE)
    emp['desc_inss_empregador'] = safe_float(m_inss_emp.group(2)) if m_inss_emp else 0.0
    emp['total_desc_inss'] = emp['desc_inss_empregador']

    emp['proventos'] = valor_resumo(block, r'Proventos')
    emp['proventos_calculados'] = emp['proventos']
    emp['base_inss'] = valor_resumo(block, r'Base INSS')
    emp['descontos'] = valor_resumo(block, r'Descontos')
    emp['excedente_inss'] = valor_resumo(block, r'Excedente INSS')
    emp['informativa'] = valor_resumo(block, r'Informativa')
    emp['base_fgts'] = valor_resumo(block, r'Base FGTS')
    emp['informativa_dedutora'] = valor_resumo(block, r'Informativa Dedutora')
    emp['valor_fgts'] = valor_resumo(block, r'Valor FGTS')
    emp['liquido'] = valor_resumo(block, r'L(?:í|Ã­)quido')
    emp['base_irrf'] = valor_resumo(block, r'Base IRRF')


# ════════════════════════════════════════════════════════════════════════════
# ─── PARSE PDF ───────────────────────────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

def extrair_rubricas_por_codigo(block):
    rubricas = {}
    padrao = r'\b(\d+)\s+([A-ZÁÉÍÓÚÂÊÔÃÕÀÜÇÇ\s\.\-\/]+?)\s+[\d:.,]+\s+([\d.,]+)\s*([PD])'
    for match in re.finditer(padrao, block, re.IGNORECASE):
        codigo = match.group(1).strip()
        descricao = match.group(2).strip()
        valor = safe_float(match.group(3))
        tipo = match.group(4).upper()
        chave = f"rubrica_{codigo}"
        if chave not in rubricas:
            rubricas[chave] = {
                'codigo': codigo,
                'descricao': descricao,
                'valor_provento': 0.0,
                'valor_desconto': 0.0
            }
        if tipo == 'P':
            rubricas[chave]['valor_provento'] = valor
        elif tipo == 'D':
            rubricas[chave]['valor_desconto'] = valor
    return rubricas


def parse_pdf(path):
    with pdfplumber.open(path) as pdf:
        full = ''.join((p.extract_text() or '') + '\n' for p in pdf.pages)

    employees = []
    block_starts = list(re.finditer(r'(?:Empr\.|Contr):\s*\d+[A-ZÁÉÍÓÚÂÊÔÃÕÀÜÇ]', full))

    for i, m in enumerate(block_starts):
        start = m.start()
        end   = block_starts[i+1].start() if i+1 < len(block_starts) else len(full)
        block = full[start:end]

        emp = {}
        emp['tipo'] = 'Contribuinte/Diretor' if block.startswith('Contr') else 'Celetista'

        m_hdr = re.match(r'(?:Empr\.|Contr):\s*(\d+)(.+?)Situação:', block, re.DOTALL)
        if not m_hdr:
            continue
        emp['id']   = m_hdr.group(1).strip()
        emp['nome'] = m_hdr.group(2).strip().replace('\n', ' ')

        m_sit = re.search(r'Situação:\s*(\S+)', block)
        emp['situacao'] = m_sit.group(1) if m_sit else ''

        m_cpf = re.search(r'CPF:\s*([\d.\-]+)', block)
        emp['cpf'] = m_cpf.group(1) if m_cpf else ''

        m_adm = re.search(r'Adm:\s*([\d/]+)', block)
        emp['admissao'] = m_adm.group(1) if m_adm else ''

        m_cc  = re.search(r'CC:\s*(\d+)', block)
        emp['cc'] = m_cc.group(1) if m_cc else ''

        m_dep = re.search(r'Depto:\s*(\d+)', block)
        emp['depto'] = m_dep.group(1) if m_dep else ''

        m_car = re.search(r'Cargo:\s*\d+(.+?)C\.B\.O', block, re.DOTALL)
        emp['cargo'] = m_car.group(1).strip().replace('\n', ' ') if m_car else ''

        m_sal = re.search(r'Salário:\s*([\d.,]+)', block)
        emp['salario_base'] = safe_float(m_sal.group(1)) if m_sal else 0.0

        m_haf = re.search(r'\b5\s+HORAS\s+AFAST\.INSS\b.*?(\d+:\d+)\s+([\d.,]+)', block, re.IGNORECASE)
        emp['horas_afastadas_inss'] = safe_float(m_haf.group(2)) if m_haf else 0.0
        m_fp = re.search(r'\b29\s*FERIAS\s+PROPORCIONAIS\b.*?([\d.,]+)\s*P', block, re.IGNORECASE | re.DOTALL)
        emp['ferias_proporcionais'] = safe_float(m_fp.group(1)) if m_fp else 0.0
        m_pl = re.search(r'\b100\s*PRO-LABORE\b.*?([\d.,]+)\s*P', block, re.IGNORECASE)
        emp['pro_labore'] = safe_float(m_pl.group(1)) if m_pl else 0.0

        m_fmh = re.search(r'\b815\s*FERIAS\s+PROP\.?MÉDIA\s+HORAS\b.*?([\d.,]+)\s*P', block, re.IGNORECASE | re.DOTALL)
        emp['ferias_media_horas'] = safe_float(m_fmh.group(1)) if m_fmh else 0.0

        m_fmv = re.search(r'\b816\s*FÉRIAS\s+PROP\.?MÉDIA\s+VALOR\b.*?([\d.,]+)\s*P', block, re.IGNORECASE | re.DOTALL)
        emp['ferias_media_valor'] = safe_float(m_fmv.group(1)) if m_fmv else 0.0

        m_ter = re.search(r'\b8169\s*1/3\s+FERIAS\s+PROPORCIONAIS\b.*?([\d.,]+)\s*P', block, re.IGNORECASE | re.DOTALL)
        emp['terco_ferias_prop'] = safe_float(m_ter.group(1)) if m_ter else 0.0

        m_13 = re.search(r'\b8550\s*13º\s+SALARIO\b.*?([\d.,]+)\s*P', block, re.IGNORECASE | re.DOTALL)
        emp['decimo_terceiro'] = safe_float(m_13.group(1)) if m_13 else 0.0

        m_ssh = re.search(r'\b9179\s*SALDO\s+DE\s+SALARIO\s+HORAS\b.*?([\d.,]+)\s*P', block, re.IGNORECASE | re.DOTALL)
        emp['saldo_salario_horas'] = safe_float(m_ssh.group(1)) if m_ssh else 0.0

        m_dff = re.search(
            r'\d*\s*DIFEREN[CÇ]A\s+DE\s+FERIAS\b.*?([\d.,]+)\s*P',
            block, re.IGNORECASE | re.DOTALL
        )
        emp['dif_ferias'] = safe_float(m_dff.group(1)) if m_dff else 0.0

        m_dft = re.search(
            r'\d*\s*DIF\.?\s*1/3\s+DE\s+FERIAS\s+[\d.,]+\s+([\d.,]+)\s*P',
            block, re.IGNORECASE
        )
        emp['dif_terco_ferias'] = safe_float(m_dft.group(1)) if m_dft else 0.0

        m_dmhf = re.search(
            r'\d*\s*DIFEREN[CÇ]A\s+M[ÉE]DIA\s+HORA\s+FERIAS\s+[\d:.,]+\s+([\d.,]+)\s*P',
            block, re.IGNORECASE
        )
        emp['dif_media_hora_ferias'] = safe_float(m_dmhf.group(1)) if m_dmhf else 0.0

        m_hn = re.search(r'HORAS\s+NORMAIS.*?(\d+:\d+)\s+([\d.,]+)', block, re.IGNORECASE)
        emp['horas_normais'] = safe_float(m_hn.group(2)) if m_hn else 0.0

        m_hf = re.search(r'HORAS\s+FERIAS.*?(\d+:\d+)\s+([\d.,]+)', block, re.IGNORECASE)
        emp['horas_ferias'] = safe_float(m_hf.group(2)) if m_hf else 0.0

        m_he50 = re.search(r'\d+HORAS EXTRAS 50%\s+[\d:]+\s+([\d.,]+)P', block)
        emp['horas_extras_50'] = safe_float(m_he50.group(1)) if m_he50 else 0.0

        m_he100 = re.search(r'\d+HORAS EXTRAS 100%\s+[\d:]+\s+([\d.,]+)P', block)
        emp['horas_extras_100'] = safe_float(m_he100.group(1)) if m_he100 else 0.0

        m_rep = re.search(r'\d+REP\.S/HORAS EXTRAS\s+[\d.,]+\s+([\d.,]+)P', block)
        emp['rep_horas_extras'] = safe_float(m_rep.group(1)) if m_rep else 0.0

        m_com = re.search(r'\d+COMISSOES\s+([\d.,]+)\s+([\d.,]+)P', block, re.IGNORECASE)
        emp['comissoes'] = safe_float(m_com.group(2)) if m_com else 0.0

        m_rep_com = re.search(r'\d+\s*REPOUSO\s+S/\s*COMISSOES\s+[\d.,]+\s+([\d.,]+)\s*P', block, re.IGNORECASE)
        emp['repouso_sem_comissoes'] = safe_float(m_rep_com.group(1)) if m_rep_com else 0.0

        if emp['nome'] and emp['nome'].strip().upper() in NOMES_ESPECIAIS_CLT:
            m_prov_especial = re.search(r'Proventos:\s*([\d.,]+)', block)
            emp['proventos_calculados'] = safe_float(m_prov_especial.group(1)) if m_prov_especial else 0.0
        else:
            emp['proventos_calculados'] = (
                emp['horas_normais']
                + emp['horas_ferias']
                + emp['horas_afastadas_inss']
                + emp['ferias_proporcionais']
                + emp['comissoes']
                + emp['pro_labore']
                + emp['horas_extras_50']
                + emp['horas_extras_100']
                + emp['rep_horas_extras']
                + emp['ferias_media_valor']
                + emp['ferias_media_horas']
                + emp['repouso_sem_comissoes']
                + emp['terco_ferias_prop']
                + emp['dif_ferias']
                + emp['dif_terco_ferias']
                + emp['dif_media_hora_ferias']
                + emp['decimo_terceiro']
                + emp['saldo_salario_horas']
                + emp['pro_labore']
            )

        # ── DESC.VALE TRANSPORTE — soma rubricas 48 e 220 ──────────────────
        m_dvt_48 = re.search(
            r'\b48\s+DESC\.VALE\s*TRANSPORTE\b(?!\s*\d{2}/\d{2})\s*[\d.,]+\s+([\d.,]+)\s*D',
            block, re.IGNORECASE
        )
        desc_vt_48 = safe_float(m_dvt_48.group(1)) if m_dvt_48 else 0.0

        m_dvt_220 = re.search(
            r'\b220\s+DESC\.VALE\s*TRANSPORTE\s*\d{2}/\d{2}\b\s*[\d.,]+\s+([\d.,]+)\s*D',
            block, re.IGNORECASE
        )
        desc_vt_220 = safe_float(m_dvt_220.group(1)) if m_dvt_220 else 0.0

        emp['desc_vale_transporte'] = desc_vt_48 + desc_vt_220

        if 'VALE TRANSPORTE' in block.upper():
            print(f"DEBUG VALE TRANSPORTE: código 48 = {desc_vt_48} | código 220 = {desc_vt_220} | total = {emp['desc_vale_transporte']}", flush=True)

        m_liqr = re.search(r'\b51\s*LIQUIDO\s+RESCISAO\b.*?([\d.,]+)\s*D', block, re.IGNORECASE | re.DOTALL)
        emp['liquido_rescisao'] = safe_float(m_liqr.group(1)) if m_liqr else 0.0

        m_inss_13resc = re.search(r'\b989\s*DESC\.INSS\s+13\s+SAL\.RESCISAO\b.*?([\d.,]+)\s*D', block, re.IGNORECASE | re.DOTALL)
        emp['desc_inss_13sal_rescisao'] = safe_float(m_inss_13resc.group(1)) if m_inss_13resc else 0.0

        m_odonto = re.search(r'\b201\s+DESC\.PLANO\s+ODONTO[\-\s]SUL\s+AMERICA\s+([\d.,]+)', block, re.IGNORECASE)
        emp['desc_plano_odonto'] = safe_float(m_odonto.group(1)) if m_odonto else 0.0

        m_saude = re.search(r'\b207\s+DESC\.PLANO\s+SAUDE[\-\s]UNIMED\s+([\d.,]+)', block, re.IGNORECASE)
        emp['desc_plano_saude'] = safe_float(m_saude.group(1)) if m_saude else 0.0

        m_comb = re.search(r'\b277\s+DESC\.AUX\.COMBUSTIVEL\s+[\d.,]+%?\s+([\d.,]+)', block, re.IGNORECASE)
        emp['desc_aux_combustivel'] = safe_float(m_comb.group(1)) if m_comb else 0.0

        m_adian = re.search(r'\b278\s*DESC\.\s*VALE\s+ADIANTAMENTO\b.*?([\d.,]+)\s*D', block, re.IGNORECASE | re.DOTALL)
        emp['desc_vale_adiantamento'] = safe_float(m_adian.group(1)) if m_adian else 0.0

        m_dha = re.search(r'\b988\s+DESCONTO\s+HORAS\s+AFASTADAS\s+[\d:.,]+\s+([\d.,]+)', block, re.IGNORECASE)
        emp['desc_horas_afastadas'] = safe_float(m_dha.group(1)) if m_dha else 0.0

        m_festa = re.search(r'\b254\s+DESC\.\s*VALE\s+FESTA\s+[\d.,]+\s+([\d.,]+)\s*D', block, re.IGNORECASE)
        emp['desc_vale_festa'] = safe_float(m_festa.group(1)) if m_festa else 0.0

        m_falt = re.search(r'\b41\s+HORAS\s+FALTAS/ATRASOS\s+[\d:.,]+\s+([\d.,]+)\s*D', block, re.IGNORECASE)
        emp['desc_faltas_atrasos'] = safe_float(m_falt.group(1)) if m_falt else 0.0

        m_desc_inss = re.search(r'\b998\s+DESC\.INSS\b.*?[\d.,]+\s+([\d.,]+)\s*D', block, re.IGNORECASE)
        emp['desc_inss'] = safe_float(m_desc_inss.group(1)) if m_desc_inss else 0.0

        m_inss_resc = re.search(r'\b826\s*DESC\.INSS\s+SOBRE\s+RESCISAO\b.*?[\d.,]+\s+([\d.,]+)\s*D', block, re.IGNORECASE)
        emp['desc_inss_sobre_rescisao'] = safe_float(m_inss_resc.group(1)) if m_inss_resc else 0.0

        m_inss_13resc = re.search(r'\b989\s+DESC\.INSS\s+13\s+SAL\.RESCISAO\b.*?[\d.,]+\s+([\d.,]+)\s*D', block, re.IGNORECASE)
        emp['desc_inss_13sal_rescisao'] = safe_float(m_inss_13resc.group(1)) if m_inss_13resc else 0.0

        m_inss_emp = re.search(r'843\s+DESC\.INSS\s+EMPREGADOR\s+[\d.,]+\s+([\d.,]+)D', block)
        emp['desc_inss_empregador'] = safe_float(m_inss_emp.group(1)) if m_inss_emp else 0.0

        m_inss_ferias = re.search(r'\bDESC\.INSS\s+FERIAS\b.*?[\d.,]+\s+([\d.,]+)\s*D', block, re.IGNORECASE)
        emp['desc_inss_ferias'] = safe_float(m_inss_ferias.group(1)) if m_inss_ferias else 0.0

        m_inss_comp = re.search(r'DESC\.COMPLEMENTO\s+INSS\s+[\d.,]+\s+([\d.,]+)D', block)
        emp['desc_complemento_inss'] = safe_float(m_inss_comp.group(1)) if m_inss_comp else 0.0

        m_irrf = re.search(r'IMPOSTO DE RENDA\s+[\d.,]+\s+([\d.,]+)D', block)
        emp['desc_irrf'] = safe_float(m_irrf.group(1)) if m_irrf else 0.0

        emp['total_desc_inss'] = (
            emp['desc_inss']
            + emp['desc_inss_sobre_rescisao']
            + emp['desc_inss_13sal_rescisao']
            + emp['desc_inss_empregador']
            + emp['desc_inss_ferias']
            + emp['desc_complemento_inss']
        )

        m_prov = re.search(r'Proventos:\s*([\d.,]+)', block)
        emp['proventos'] = safe_float(m_prov.group(1)) if m_prov else 0.0

        m_base_inss = re.search(r'Base INSS:\s*([\d.,]+)', block)
        emp['base_inss'] = safe_float(m_base_inss.group(1)) if m_base_inss else 0.0

        m_desc = re.search(r'Descontos:\s*([\d.,]+)', block)
        emp['descontos'] = safe_float(m_desc.group(1)) if m_desc else 0.0

        m_exc_inss = re.search(r'Excedente INSS:\s*([\d.,]+)', block)
        emp['excedente_inss'] = safe_float(m_exc_inss.group(1)) if m_exc_inss else 0.0

        m_info = re.search(r'Informativa:\s*([\d.,]+)', block)
        emp['informativa'] = safe_float(m_info.group(1)) if m_info else 0.0

        m_base_fgts = re.search(r'Base FGTS:\s*([\d.,]+)', block)
        emp['base_fgts'] = safe_float(m_base_fgts.group(1)) if m_base_fgts else 0.0

        m_info_ded = re.search(r'Informativa Dedutora:\s*([\d.,]+)', block)
        emp['informativa_dedutora'] = safe_float(m_info_ded.group(1)) if m_info_ded else 0.0

        m_valor_fgts = re.search(r'Valor\s+FGTS:\s*([\d.,]+)', block, re.IGNORECASE)
        emp['valor_fgts'] = safe_float(m_valor_fgts.group(1)) if m_valor_fgts else 0.0

        m_liquido = re.search(r'Líquido:\s*([\d.,]+)', block)
        emp['liquido'] = safe_float(m_liquido.group(1)) if m_liquido else 0.0

        m_base_irrf = re.search(r'Base IRRF:\s*([-\d.,]+)', block)
        emp['base_irrf'] = safe_float(m_base_irrf.group(1)) if m_base_irrf else 0.0

        base_ind = emp['base_fgts']
        emp['base_encargos'] = base_ind
        emp['rat_valor']       = round(base_ind * 0.005, 2)
        emp['empresa_valor']   = round(base_ind * 0.20,  2)
        emp['terceiros_valor'] = round(base_ind * 0.058, 2)

        obs_partes = []

        m_dem = re.search(r'DEMITIDO EM\s*([\d/]+)\s*-\s*MOTIVO\s+(.+?)$', block, re.MULTILINE)
        if m_dem:
            obs_partes.append(f"Demitido em {m_dem.group(1)} — {m_dem.group(2).strip()}")

        m_fer_periodo = re.search(r'FERIAS\s+DE\s+([\d/]+)\s*-\s*([\d/]+)', block, re.IGNORECASE)
        if m_fer_periodo:
            emp['ferias_periodo_inicio'] = m_fer_periodo.group(1)
            emp['ferias_periodo_fim']    = m_fer_periodo.group(2)
            obs_partes.append(f"Férias de {m_fer_periodo.group(1)} a {m_fer_periodo.group(2)}")
        else:
            emp['ferias_periodo_inicio'] = ''
            emp['ferias_periodo_fim']    = ''

        m_doenca = re.search(r'Doen[çc]a\s+per[íi]odo\s+superior\s+a\s+15\s+dias:\s*([\d/]+)\s*a\s*([\d/]+)', block, re.IGNORECASE)
        if m_doenca:
            emp['doenca_inicio'] = m_doenca.group(1)
            emp['doenca_fim']    = m_doenca.group(2)
            obs_partes.append(f"Afastado(a) por doença desde {m_doenca.group(1)} (período superior a 15 dias)")
        else:
            emp['doenca_inicio'] = ''
            emp['doenca_fim']    = ''

        emp['obs'] = ' | '.join(obs_partes)

        emp['rubricas'] = extrair_rubricas_por_codigo(block)

        aplicar_regra_fabiano(emp, block)

        employees.append(emp)

    encargos          = _parse_encargos(full)
    enc_rat_consolidado = _parse_rat_empresa_terceiros(full)

    return employees, encargos, enc_rat_consolidado


def _parse_encargos(full):
    encargos = []
    apuracao_patterns = [
        ('INSS Segurado (Folha)',      r'INSS Segurado\(Folha\):\s*([\d.,]+)(?:.*?)([\d.,]+)\s*$'),
        ('INSS Empresa e RAT (Folha)', r'INSS Empresa e RAT\(Folha\):\s*([\d.,]+)(?:.*?)([\d.,]+)\s*$'),
        ('INSS Terceiros (Folha)',     r'INSS Terceiros\(Folha\):\s*([\d.,]+)(?:.*?)([\d.,]+)\s*$'),
        ('IRRF (Folha)',               r'IRRF\(Folha\):\s*([\d.,]+)(?:.*?)([\d.,]+)\s*$'),
    ]
    for nome, pat in apuracao_patterns:
        m = re.search(pat, full, re.MULTILINE)
        if m:
            valor    = safe_float(m.group(1))
            recolher = safe_float(m.group(2))
            if 'Segurado' in nome:
                mb = re.search(r'Salário contribuição empregados:\s*([\d.,]+)', full)
                base = safe_float(mb.group(1)) if mb else valor
            elif 'Empresa' in nome or 'Terceiros' in nome:
                mb = re.search(r'Base total:\s*([\d.,]+)', full)
                base = safe_float(mb.group(1)) if mb else valor
            elif 'IRRF' in nome:
                bases = re.findall(r'Base IRRF Mensal:\s*([\d.,]+)', full)
                base = safe_float(bases[-1]) if bases else valor
            else:
                base = valor
            encargos.append((nome, base, recolher))
    mb_fgts = re.search(r'Base do FGTS:\s*([\d.,]+)', full)
    mv_fgts = re.search(r'Valor do FGTS:\s*([\d.,]+)', full)
    encargos.append(('FGTS (Folha)',
                     safe_float(mb_fgts.group(1)) if mb_fgts else 0.0,
                     safe_float(mv_fgts.group(1)) if mv_fgts else 0.0))
    return encargos


def _parse_rat_empresa_terceiros(full):
    m_sec = re.search(r'Resumo por Rubrica(.+?)INSS FGTS', full, re.DOTALL)
    secao = m_sec.group(1) if m_sec else full

    rubricas = [
        ('Horas Normais',              r'\d+HORAS NORMAIS\s+[\d:,.]+\s+([\d.,]+)P'),
        ('Horas Ferias',               r'\d+HORAS FERIAS\s+[\d:,.]+\s+([\d.,]+)P'),
        ('Comissões',                  r'\d+COMISSOES\s+[\d.,]+\s+([\d.,]+)P'),
        ('Horas Extras 50%',           r'\d+HORAS EXTRAS 50%\s+[\d:,.]+\s+([\d.,]+)P'),
        ('Rep. S/ Horas Extras',       r'\d+REP\.S/HORAS EXTRAS\s+[\d.,]+\s+([\d.,]+)P'),
        ('Férias Proporcionais',       r'\d+\s*FERIAS\s+PROPORCIONAIS\s+[\d.,]+\s+([\d.,]+)P'),
        ('Férias Méd. Horas',          r'\d+\s*F[ÉE]RIAS\s+(?:PROP\.?\s*)?M[ÉE]DIA\s+HORAS\s+[\d.,]+\s+([\d.,]+)P'),
        ('Férias Méd. Valor',          r'\d+\s*F[ÉE]RIAS\s+(?:PROP\.?\s*)?M[ÉE]DIA\s+VALOR\s+[\d.,]+\s+([\d.,]+)P'),
        ('Repouso S/ Comissões',       r'\d+REPOUSO S/ COMISSOES\s+[\d.,]+\s+([\d.,]+)P'),
        ('1/3 Férias Proporcionais',   r'\d+\s*1/3\s+(?:DAS\s+)?FERIAS(?:\s+PROPORCIONAIS)?\s+[\d.,]+\s+([\d.,]+)P'),
        ('13º Salário',                r'\d+13º SALARIO\s+[\d.,]+\s+([\d.,]+)P'),
        ('Saldo de Salário Horas',     r'\d+SALDO DE SALARIO HORAS\s+[\d:,.]+\s+([\d.,]+)P'),
        ('Pró-Labore',                 r'\d+PRO-LABORE\s+[\d:,.]+\s+([\d.,]+)P'),
        ('Diferença de Férias',        r'DIFEREN[CÇ]A\s+DE\s+FERIAS\s+[\d.,]+\s+([\d.,]+)P'),
        ('Dif. 1/3 de Férias',         r'DIF\.?\s*1/3\s+DE\s+FERIAS\s+[\d.,]+\s+([\d.,]+)P'),
        ('Diferença Média Hora Férias',r'DIFEREN[CÇ]A\s+M[ÉE]DIA\s+HORA\s+FERIAS\s+[\d.,]+\s+([\d.,]+)P'),
    ]

    detalhes = []
    total_base = 0.0
    total_sem_prolabore = 0.0

    for nome, pat in rubricas:
        m = re.search(pat, secao, re.IGNORECASE)
        valor = safe_float(m.group(1)) if m else 0.0
        if valor > 0:
            detalhes.append((nome, valor))
            total_base += valor
            if nome != 'Pró-Labore':
                total_sem_prolabore += valor

    return {
        'detalhes':        detalhes,
        'base':            total_sem_prolabore,
        'base_empresa':    total_base,
        'rat_aliq':        0.005,
        'empresa_aliq':    0.20,
        'terceiros_aliq':  0.058,
        'rat_valor':       round(total_sem_prolabore * 0.005,  2),
        'empresa_valor':   round(total_base          * 0.20,   2),
        'terceiros_valor': round(total_sem_prolabore * 0.058,  2),
    }


# ════════════════════════════════════════════════════════════════════════════
# ─── EXCEL PRINCIPAL ─────────────────────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

def create_excel(employees, encargos, enc_rat):
    wb = Workbook()

    NAVY     = '0D1B2A'
    MID_NAVY = '1B3A5C'
    BLUE     = '2176FF'
    LT_BLUE  = 'D6E8FF'
    WHITE    = 'FFFFFF'
    GRAY_BG  = 'F4F7FB'
    GRAY_BD  = 'CBD5E1'
    RED      = 'C0392B'
    GOLD     = 'D97706'
    GREEN    = '1B5E20'
    LT_GREEN = 'E8F5E9'
    TEAL     = '004D61'
    LT_TEAL  = 'E0F4F8'
    ORANGE   = 'E65100'
    LT_ORNG  = 'FFF3E0'
    PURPLE   = '4A148C'
    LT_PURP  = 'F3E5F5'

    def fill(c):  return PatternFill('solid', fgColor=c)
    def font(bold=False, color=None, size=10, name='Calibri'):
        return Font(name=name, bold=bold, color=color or '000000', size=size)
    thin = Side(style='thin',   color=GRAY_BD)
    med  = Side(style='medium', color=BLUE)
    def border(bottom=thin): return Border(left=thin, right=thin, top=thin, bottom=bottom)

    c_alt = fill(GRAY_BG); c_wht = fill(WHITE)
    ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lft = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    rgt = Alignment(horizontal='right',  vertical='center')
    mny = 'R$ #,##0.00'
    pct = '0.00%'

    DEPTO_NAMES = {
        '1': 'Comercial / Adm', '2': 'Estoque / Logística',
        '3': 'TI', '4': 'Financeiro', '5': 'Marketing'
    }

    def titulo_aba(ws, txt, subtxt, ncols):
        col_end = get_column_letter(ncols)
        ws.merge_cells(f'A1:{col_end}1')
        c = ws['A1']
        c.value = txt; c.font = font(True, WHITE, 13)
        c.fill = fill(BLUE); c.alignment = ctr
        ws.row_dimensions[1].height = 32
        ws.merge_cells(f'A2:{col_end}2')
        c = ws['A2']
        c.value = subtxt; c.font = font(False, WHITE, 9)
        c.fill = fill(MID_NAVY); c.alignment = ctr
        ws.row_dimensions[2].height = 18
        ws.row_dimensions[3].height = 10

    def secao(ws, row, txt, ncols, cor=TEAL):
        col_end = get_column_letter(ncols)
        ws.merge_cells(f'A{row}:{col_end}{row}')
        c = ws.cell(row=row, column=1, value=txt)
        c.font = font(True, WHITE, 11); c.fill = fill(cor); c.alignment = ctr
        ws.row_dimensions[row].height = 24

    def cabecalho(ws, row, headers_widths):
        for ci, (lbl, w) in enumerate(headers_widths, 1):
            c = ws.cell(row=row, column=ci, value=lbl)
            c.font = font(True, WHITE, 9); c.fill = fill(MID_NAVY)
            c.alignment = ctr; c.border = border()
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[row].height = 26

    def linha_total(ws, row, label, col_soma_inicio, col_soma_fim, ncols_merge=None, cor=LT_BLUE):
        nm = ncols_merge or col_soma_inicio - 1
        ws.merge_cells(f'A{row}:{get_column_letter(nm)}{row}')
        c = ws.cell(row=row, column=1, value=label)
        c.font = font(True, NAVY, 10); c.fill = fill(cor); c.alignment = ctr; c.border = border(med)
        for ci in range(col_soma_inicio, col_soma_fim + 1):
            cl = get_column_letter(ci)
            c = ws.cell(row=row, column=ci, value=f'=SUM({cl}8:{cl}{row-1})')
            c.font = font(True, NAVY, 10); c.fill = fill(cor)
            c.alignment = rgt; c.border = border(med); c.number_format = mny
        ws.row_dimensions[row].height = 24

    ws = wb.active
    ws.title = 'Resumo Geral'
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A8'

    ws.merge_cells('A1:AC1')
    c = ws['A1']
    c.value = '📊  FOLHA DE PAGAMENTO DETALHADA — COMPETÊNCIA 2026'
    c.font = font(True, WHITE, 13); c.fill = fill(BLUE); c.alignment = ctr
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:AC2')
    c = ws['A2']
    c.value = '3I IMPORTAÇÃO E EXPORTAÇÃO LTDA  •  CNPJ: 20.783.843/0001-19  •  Emissão: 06/05/2026'
    c.font = font(True, WHITE, 10); c.fill = fill(MID_NAVY); c.alignment = ctr
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 8

    cols_rg = [
    ('ID', 6), ('Nome', 30), ('Tipo', 12), ('Cargo', 20),
    ('Situação', 12), ('CPF', 14), ('Admissão', 11),
    ('Hs Normais', 12), ('Horas Ferias', 12), ('Hs Afastadas INSS', 14), ('Férias Prop.', 13),
    ('Comissões', 14), ('Rep.S/Com.', 14),
    ('Pró-Labore', 13), ('HE 50%', 12), ('Rep.S/HE', 12),
    ('Férias Méd.Hs', 13), ('Férias Méd.Val', 13),
    ('1/3 Fér.Prop.', 13), ('13º Salário', 13), ('Saldo Sal.Hs', 13),
    ('Dif.Férias', 13), ('Dif.1/3 Fér.', 13), ('Dif.Méd.Hs Fér.', 14),
    ('Proventos Calculados (R$)', 22),
    ('Desc.VT', 12), ('Liq.Rescisão', 13),
    ('Desc.Odonto', 13), ('Desc.Saúde', 13),
    ('Desc.Comb.1%', 13), ('Desc.Adiant.', 13), ('Desc.Vale Festa', 14), ('Desc.Faltas/Atrasos', 16),
    ('Desc.INSS', 12),
    ('Desc.INSS Resc.', 13), ('Desc.INSS 13ºResc.', 15), ('Desc.INSS Empregador', 16),
    ('Desc.INSS Férias', 15), ('Desc.Compl.INSS', 15),
    ('Desc.IRRF', 12), ('Desc.Hs Afastadas', 15),
    ('Proventos (R$)', 14), ('Base INSS (R$)', 14), ('Descontos (R$)', 14),
    ('Excedente INSS', 13), ('Informativa', 12), ('Base FGTS (R$)', 14),
    ('Valor FGTS (R$)', 14), ('Líquido (R$)', 14), ('Base IRRF (R$)', 14), ('Obs', 24),
]
    for i, (lbl, w) in enumerate(cols_rg, 1):
        c = ws.cell(row=4, column=i, value=lbl)
        c.font = font(True, WHITE, 8)
        c.fill = fill(PURPLE) if lbl == 'Proventos Calculados (R$)' else fill(MID_NAVY)
        c.alignment = ctr; c.border = border()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 30

    ncols_rg = len(cols_rg)
    col_prov_calc = next(i+1 for i, (lbl, _) in enumerate(cols_rg) if lbl == 'Proventos Calculados (R$)')

    for ri, emp in enumerate(employees, 5):
        sit = emp.get('situacao', '').lower()
        row_fill = (fill('FFF3CD') if sit == 'demitido'
                    else fill('FFE4E4') if sit in ('doença', 'afastado')
                    else c_alt if ri % 2 == 0 else c_wht)
        vals = [
    emp['id'], emp['nome'], emp['tipo'], emp['cargo'],
    emp['situacao'], emp['cpf'], emp['admissao'],
    emp['horas_normais'], emp['horas_ferias'], emp['horas_afastadas_inss'], emp['ferias_proporcionais'],
    emp['comissoes'], emp['repouso_sem_comissoes'],
    emp['pro_labore'], emp['horas_extras_50'], emp['rep_horas_extras'],
    emp['ferias_media_horas'], emp['ferias_media_valor'],
    emp['terco_ferias_prop'], emp['decimo_terceiro'], emp['saldo_salario_horas'],
    emp.get('dif_ferias', 0.0), emp.get('dif_terco_ferias', 0.0), emp.get('dif_media_hora_ferias', 0.0),
    emp['proventos_calculados'],
    emp['desc_vale_transporte'], emp['liquido_rescisao'],
    emp['desc_plano_odonto'], emp['desc_plano_saude'],
    emp['desc_aux_combustivel'], emp['desc_vale_adiantamento'], emp.get('desc_vale_festa', 0.0), emp.get('desc_faltas_atrasos', 0.0),
    emp['desc_inss'],
    emp['desc_inss_sobre_rescisao'], emp['desc_inss_13sal_rescisao'], emp['desc_inss_empregador'],
    emp['desc_inss_ferias'], emp['desc_complemento_inss'],
    emp['desc_irrf'], emp['desc_horas_afastadas'],
    emp['proventos'], emp['base_inss'], emp['descontos'],
    emp['excedente_inss'], emp['informativa'], emp['base_fgts'],
    emp['valor_fgts'], emp['liquido'], emp['base_irrf'],
    emp['obs'],
]
        num_cols = set(range(8, ncols_rg))
        for ci, v in enumerate(vals, 1):
            fnt = font(True, MID_NAVY, 8) if ci == 2 else font(size=8)
            if ci == 5:
                sit_color = {'demitido': RED, 'doença': GOLD, 'afastado': GOLD}.get(sit)
                if sit_color: fnt = font(True, sit_color, 8)
            if ci == col_prov_calc:
                fnt = font(True, PURPLE, 9)
            aln = ctr if ci in (1, 3, 5, 6, 7) else (rgt if ci in num_cols else lft)
            fmt = mny if ci in num_cols else None
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = fnt
            c.fill = fill(LT_PURP) if ci == col_prov_calc else row_fill
            c.alignment = aln; c.border = border()
            if fmt: c.number_format = fmt
        ws.row_dimensions[ri].height = 18

    tr = len(employees) + 5
    ws.merge_cells(f'A{tr}:G{tr}')
    c = ws.cell(row=tr, column=1, value='TOTAIS GERAIS')
    c.font = font(True, NAVY, 10); c.fill = fill(LT_BLUE); c.alignment = ctr; c.border = border(med)
    for ci in range(8, ncols_rg):
        cl = get_column_letter(ci)
        c = ws.cell(row=tr, column=ci, value=f'=SUM({cl}5:{cl}{tr-1})')
        c.font = font(True, PURPLE if ci == col_prov_calc else NAVY, 10)
        c.fill = fill(LT_PURP) if ci == col_prov_calc else fill(LT_BLUE)
        c.alignment = rgt; c.border = border(med); c.number_format = mny
    ws.row_dimensions[tr].height = 24

    ws2 = wb.create_sheet('Por Departamento')
    ws2.sheet_view.showGridLines = False
    titulo_aba(ws2, '🏢  RESUMO POR DEPARTAMENTO', '3I IMPORTAÇÃO E EXPORTAÇÃO LTDA', 6)

    cabecalho(ws2, 4, [('Depto', 8), ('Descrição', 28), ('Funcionários', 14),
                        ('Proventos (R$)', 18), ('Descontos (R$)', 18), ('Líquido (R$)', 18)])
    deptos = {}
    for e in employees:
        deptos.setdefault(e.get('depto', '?'), []).append(e)
    row = 5
    for dk in sorted(deptos):
        emps = deptos[dk]
        prov = sum(e['proventos'] for e in emps)
        desc = sum(e['descontos'] for e in emps)
        liq  = sum(e['liquido']   for e in emps)
        rf   = c_alt if row % 2 == 0 else c_wht
        for ci, v in enumerate([dk, DEPTO_NAMES.get(dk, f'Depto {dk}'), len(emps), prov, desc, liq], 1):
            c = ws2.cell(row=row, column=ci, value=v)
            c.fill = rf; c.border = border(); c.font = font()
            c.alignment = ctr if ci <= 3 else rgt
            if ci > 3: c.number_format = mny
        ws2.row_dimensions[row].height = 20; row += 1
    ws2.merge_cells(f'A{row}:B{row}')
    c = ws2.cell(row=row, column=1, value='TOTAL GERAL')
    c.font = font(True, NAVY, 10); c.fill = fill(LT_BLUE); c.alignment = ctr; c.border = border(med)
    ws2.cell(row=row, column=3, value=len(employees)).font = font(True, NAVY, 10)
    ws2.cell(row=row, column=3).fill = fill(LT_BLUE); ws2.cell(row=row, column=3).alignment = ctr
    ws2.cell(row=row, column=3).border = border(med)
    for ci in range(4, 7):
        cl = get_column_letter(ci)
        c = ws2.cell(row=row, column=ci, value=f'=SUM({cl}5:{cl}{row-1})')
        c.font = font(True, NAVY, 10); c.fill = fill(LT_BLUE)
        c.alignment = rgt; c.border = border(med); c.number_format = mny
    ws2.row_dimensions[row].height = 24

    ws3 = wb.create_sheet('Encargos e Tributos')
    ws3.sheet_view.showGridLines = False
    titulo_aba(ws3, '🧾  ENCARGOS E TRIBUTOS', '3I IMPORTAÇÃO E EXPORTAÇÃO LTDA', 3)

    cabecalho(ws3, 4, [('Encargo', 34), ('Base de Cálculo (R$)', 22), ('Valor a Recolher (R$)', 22)])
    for ri, (nome, base, val) in enumerate(encargos, 5):
        rf = c_alt if ri % 2 == 0 else c_wht
        for ci, v in enumerate([nome, base, val], 1):
            c = ws3.cell(row=ri, column=ci, value=v)
            c.fill = rf; c.border = border(); c.font = font()
            c.alignment = lft if ci == 1 else rgt
            if ci > 1: c.number_format = mny
        ws3.row_dimensions[ri].height = 20
    tr3 = len(encargos) + 5
    c = ws3.cell(row=tr3, column=1, value='TOTAL A RECOLHER')
    c.font = font(True, NAVY, 10); c.fill = fill(LT_BLUE); c.alignment = lft; c.border = border(med)
    for ci in [2, 3]:
        cl = get_column_letter(ci)
        c = ws3.cell(row=tr3, column=ci, value=f'=SUM({cl}5:{cl}{tr3-1})')
        c.font = font(True, NAVY, 10); c.fill = fill(LT_BLUE)
        c.alignment = rgt; c.border = border(med); c.number_format = mny
    ws3.row_dimensions[tr3].height = 24

    ws4 = wb.create_sheet('RAT, Empresa e Terceiros')
    ws4.sheet_view.showGridLines = False
    for i, w in enumerate([34, 22, 14, 22, 22], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    titulo_aba(ws4, '🏛️  APURAÇÃO CONSOLIDADA: RAT — EMPRESA — TERCEIROS',
               'Base: proventos CLTs (excl. Pró-Labore p/ RAT e Terceiros)', 5)

    secao(ws4, 4, '📋  COMPOSIÇÃO DA BASE DE CÁLCULO', 5)
    for ci, (lbl, w) in enumerate([('Rubrica', 34), ('Valor (R$)', 22)], 1):
        c = ws4.cell(row=5, column=ci, value=lbl)
        c.font = font(True, WHITE, 10); c.fill = fill(MID_NAVY)
        c.alignment = ctr; c.border = border()
    ws4.row_dimensions[5].height = 22

    det = enc_rat['detalhes']
    for ri, (nr, vr) in enumerate(det, 6):
        rf = c_alt if ri % 2 == 0 else c_wht
        c = ws4.cell(row=ri, column=1, value=nr)
        c.fill = rf; c.border = border(); c.font = font(size=10); c.alignment = lft
        c = ws4.cell(row=ri, column=2, value=vr)
        c.fill = rf; c.border = border(); c.font = font(size=10)
        c.alignment = rgt; c.number_format = mny
        ws4.row_dimensions[ri].height = 18

    tb = len(det) + 6
    ws4.merge_cells(f'A{tb}:A{tb}')
    c = ws4.cell(row=tb, column=1, value='BASE TOTAL (sem Pró-Labore) — RAT e Terceiros')
    c.font = font(True, TEAL, 10); c.fill = fill(LT_TEAL); c.alignment = lft; c.border = border(med)
    c = ws4.cell(row=tb, column=2, value=enc_rat['base'])
    c.font = font(True, TEAL, 10); c.fill = fill(LT_TEAL)
    c.alignment = rgt; c.border = border(med); c.number_format = mny
    ws4.row_dimensions[tb].height = 22

    tb2 = tb + 1
    ws4.merge_cells(f'A{tb2}:A{tb2}')
    c = ws4.cell(row=tb2, column=1, value='BASE TOTAL (com Pró-Labore) — Empresa')
    c.font = font(True, GREEN, 10); c.fill = fill(LT_GREEN); c.alignment = lft; c.border = border(med)
    c = ws4.cell(row=tb2, column=2, value=enc_rat['base_empresa'])
    c.font = font(True, GREEN, 10); c.fill = fill(LT_GREEN)
    c.alignment = rgt; c.border = border(med); c.number_format = mny
    ws4.row_dimensions[tb2].height = 22
    ws4.row_dimensions[tb2 + 1].height = 12

    re_row = tb2 + 2
    secao(ws4, re_row, '💰  APURAÇÃO DOS ENCARGOS', 5)
    re_row += 1
    for ci, lbl in enumerate(['Encargo', 'Base de Cálculo (R$)', 'Alíquota', 'Valor Calculado (R$)', 'Observação'], 1):
        c = ws4.cell(row=re_row, column=ci, value=lbl)
        c.font = font(True, WHITE, 10); c.fill = fill(MID_NAVY)
        c.alignment = ctr; c.border = border()
    ws4.row_dimensions[re_row].height = 22

    linhas_enc = [
        ('RAT  (Riscos Ambientais do Trabalho)',
         enc_rat['base'], enc_rat['rat_aliq'], enc_rat['rat_valor'],
         '0,50% — sem Pró-Labore', fill(LT_TEAL), TEAL),
        ('Empresa  (INSS Patronal)',
         enc_rat['base_empresa'], enc_rat['empresa_aliq'], enc_rat['empresa_valor'],
         '20,00% — com Pró-Labore', fill(LT_GREEN), GREEN),
        ('Terceiros  (Sistema S + Sal. Educação)',
         enc_rat['base'], enc_rat['terceiros_aliq'], enc_rat['terceiros_valor'],
         '5,80% — sem Pró-Labore', fill(LT_ORNG), ORANGE),
    ]
    for off, (en, bv, av, cv, ov, rf2, tc) in enumerate(linhas_enc):
        ri2 = re_row + 1 + off
        for ci, v in enumerate([en, bv, av, cv, ov], 1):
            c = ws4.cell(row=ri2, column=ci, value=v)
            c.fill = rf2; c.border = border(); c.font = font(True, tc, 10)
            if ci == 1:   c.alignment = lft
            elif ci == 3: c.alignment = ctr; c.number_format = pct
            elif ci == 5: c.alignment = ctr
            else:         c.alignment = rgt; c.number_format = mny
        ws4.row_dimensions[ri2].height = 22

    row_tot4 = re_row + 1 + len(linhas_enc)
    ws4.merge_cells(f'A{row_tot4}:C{row_tot4}')
    c = ws4.cell(row=row_tot4, column=1, value='TOTAL DOS ENCARGOS')
    c.font = font(True, NAVY, 11); c.fill = fill(LT_BLUE); c.alignment = ctr; c.border = border(med)
    total4 = enc_rat['rat_valor'] + enc_rat['empresa_valor'] + enc_rat['terceiros_valor']
    c = ws4.cell(row=row_tot4, column=4, value=total4)
    c.font = font(True, NAVY, 11); c.fill = fill(LT_BLUE)
    c.alignment = rgt; c.border = border(med); c.number_format = mny
    c = ws4.cell(row=row_tot4, column=5); c.fill = fill(LT_BLUE); c.border = border(med)
    ws4.row_dimensions[row_tot4].height = 26

    ws5 = wb.create_sheet('Encargos por Funcionário')
    ws5.sheet_view.showGridLines = False
    for i, w in enumerate([6, 30, 12, 22, 22, 14, 14, 14, 14], 1):
        ws5.column_dimensions[get_column_letter(i)].width = w

    titulo_aba(ws5, '👤  ENCARGOS PATRONAIS POR FUNCIONÁRIO (CLTs)',
               'RAT 0,50%  |  Empresa 20,00%  |  Terceiros 5,80%  —  Base = proventos computáveis', 9)

    cabecalho(ws5, 4, [
        ('ID', 6), ('Nome', 30), ('Cargo', 22),
        ('Base Encargos (R$)', 22), ('RAT 0,50% (R$)', 14),
        ('Empresa 20% (R$)', 14), ('Terceiros 5,80% (R$)', 14), ('Total Encargos (R$)', 16),
        ('Situação', 12),
    ])

    clts = [e for e in employees if e['tipo'] == 'Celetista']
    for ri, emp in enumerate(clts, 5):
        sit = emp.get('situacao', '').lower()
        row_fill5 = (fill('FFF3CD') if sit == 'demitido'
                     else fill('FFE4E4') if sit in ('doença', 'afastado')
                     else c_alt if ri % 2 == 0 else c_wht)
        total_enc5 = emp['rat_valor'] + emp['empresa_valor'] + emp['terceiros_valor']
        vals5 = [emp['id'], emp['nome'], emp['cargo'],
                 emp['base_encargos'], emp['rat_valor'],
                 emp['empresa_valor'], emp['terceiros_valor'], total_enc5,
                 emp['situacao']]
        for ci, v in enumerate(vals5, 1):
            c = ws5.cell(row=ri, column=ci, value=v)
            c.fill = row_fill5; c.border = border()
            c.font = font(True, MID_NAVY, 9) if ci == 2 else font(size=9)
            if ci in (1, 3, 9): c.alignment = ctr
            elif ci >= 4 and ci <= 8: c.alignment = rgt; c.number_format = mny
            else: c.alignment = lft
        ws5.row_dimensions[ri].height = 18

    tr5 = len(clts) + 5
    ws5.merge_cells(f'A{tr5}:C{tr5}')
    c = ws5.cell(row=tr5, column=1, value='TOTAIS')
    c.font = font(True, NAVY, 10); c.fill = fill(LT_BLUE); c.alignment = ctr; c.border = border(med)
    for ci in range(4, 9):
        cl = get_column_letter(ci)
        c = ws5.cell(row=tr5, column=ci, value=f'=SUM({cl}5:{cl}{tr5-1})')
        c.font = font(True, NAVY, 10); c.fill = fill(LT_BLUE)
        c.alignment = rgt; c.border = border(med); c.number_format = mny
    ws5.row_dimensions[tr5].height = 24

    ws6 = wb.create_sheet('Descontos INSS e IRRF')
    ws6.sheet_view.showGridLines = False
    for i, w in enumerate([6, 30, 12, 16, 16, 16, 16, 16, 16, 14, 16, 16, 15, 17, 14, 16, 16], 1):
        ws6.column_dimensions[get_column_letter(i)].width = w

    titulo_aba(ws6,
               '🔐  DESCONTOS DE INSS E IRRF POR FUNCIONÁRIO',
               'DESC.INSS  |  INSS s/ Rescisão  |  INSS 13º Rescisão  |  INSS Empregador  |  INSS Férias  |  Compl. INSS  |  IRRF',
               11)

    cabecalho(ws6, 4, [
    ('ID', 6), ('Nome', 30), ('Tipo', 12),
    ('Desc. INSS Normal (R$)', 16),
    ('INSS s/ Rescisão (R$)', 16),
    ('INSS 13º Rescisão (R$)', 16),
    ('INSS Empregador (R$)', 16),
    ('INSS Férias (R$)', 16),
    ('Compl. INSS (R$)', 16),
    ('Desc. IRRF (R$)', 14),
    ('Desc. Odonto (R$)', 14),
    ('Desc. Saúde (R$)', 14),
    ('Desc. VT (R$)', 13),
    ('Desc. Adiant. (R$)', 14),
    ('Desc. Comb. (R$)', 13),
    ('Desc. Vale Festa (R$)', 15),
    ('Desc. Faltas/Atrasos (R$)', 17),
    ('Hs Afastadas (R$)', 14),
    ('Total INSS (R$)', 16),
    ('Base INSS (R$)', 16),
])

    for ri, emp in enumerate(employees, 5):
        sit = emp.get('situacao', '').lower()
        row_fill6 = (fill('FFF3CD') if sit == 'demitido'
                     else fill('FFE4E4') if sit in ('doença', 'afastado')
                     else c_alt if ri % 2 == 0 else c_wht)
        vals6 = [
    emp['id'], emp['nome'], emp['tipo'],
    emp['desc_inss'],
    emp['desc_inss_sobre_rescisao'],
    emp['desc_inss_13sal_rescisao'],
    emp['desc_inss_empregador'],
    emp['desc_inss_ferias'],
    emp['desc_complemento_inss'],
    emp['desc_irrf'],
    emp['desc_plano_odonto'],
    emp['desc_plano_saude'],
    emp['desc_vale_transporte'],
    emp['desc_vale_adiantamento'],
    emp['desc_aux_combustivel'],
    emp.get('desc_vale_festa', 0.0),
    emp.get('desc_faltas_atrasos', 0.0),
    emp['desc_horas_afastadas'],
    emp['total_desc_inss'],
    emp['base_inss'],
]
        for ci, v in enumerate(vals6, 1):
            c = ws6.cell(row=ri, column=ci, value=v)
            c.fill = row_fill6; c.border = border()
            c.font = font(True, MID_NAVY, 9) if ci == 2 else font(size=9)
            if ci in (1, 3): c.alignment = ctr
            elif ci >= 4: c.alignment = rgt; c.number_format = mny
            else: c.alignment = lft
        ws6.row_dimensions[ri].height = 18

    tr6 = len(employees) + 5
    ws6.merge_cells(f'A{tr6}:C{tr6}')
    c = ws6.cell(row=tr6, column=1, value='TOTAIS')
    c.font = font(True, NAVY, 10); c.fill = fill(LT_BLUE); c.alignment = ctr; c.border = border(med)
    for ci in range(4, 21):
        cl = get_column_letter(ci)
        c = ws6.cell(row=tr6, column=ci, value=f'=SUM({cl}5:{cl}{tr6-1})')
        c.font = font(True, NAVY, 10); c.fill = fill(LT_BLUE)
        c.alignment = rgt; c.border = border(med); c.number_format = mny
    ws6.row_dimensions[tr6].height = 24

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name); tmp.close()
    return tmp.name


# ════════════════════════════════════════════════════════════════════════════
# ─── TEMPLATE DE LANÇAMENTO CONTÁBIL ─────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

LINHAS_TEMPLATE = [
    ('GL', '4.01.01.01.12', 'COMISSÕES',                    'comissoes',              None),
    ('GL', '4.01.01.06.10', 'REPOUSO S/ COMISSÕES',         'repouso_sem_comissoes',  None),
    ('GL', '4.01.01.01.01', 'PRO-LABORE',                   'proventos_calculados',   None),
    ('GL', '4.01.01.01.02', 'HORAS NORMAIS',                'horas_normais',          None),
    ('GL', '4.01.01.01.10', 'HORAS EXTRAS 50%',             'horas_extras_50',        None),
    ('GL', '4.01.01.01.44', 'LIQUIDO RESCISAO',              None, 'liquido_rescisao'),
    ('GL', '4.01.01.01.21', 'REPOUSO S/ HORAS EXTRAS',      'rep_horas_extras',       None),
    ('GL', '4.01.01.01.35', 'SALDO DE SALARIO HORAS',       'saldo_salario_horas',    None),
    ('GL', '4.01.01.01.09', 'HORAS FERIAS',                 'horas_ferias',           None),
    ('GL', '4.01.01.01.22', 'FERIAS VENCIDAS',              'ferias_vencidas',        None),
    ('GL', '4.01.01.01.23', 'FERIAS PROPORCIONAIS',         'ferias_proporcionais',   None),
    ('GL', '4.01.01.01.25', 'FERIAS MEDIA HORAS',           'ferias_media_horas',     None),
    ('GL', '4.01.01.01.26', 'FERIAS INDENIZADAS',           'ferias_indenizadas',     None),
    ('GL', '4.01.01.01.27', 'FÉRIAS PROP.MÉDIA VALOR',      'ferias_media_valor',     None),
    ('GL', '4.01.01.01.28', 'FERIAS VENC.MEDIA VALOR',      'ferias_vencidas_media_valor', None),
    ('GL', '4.01.01.01.29', 'FERIAS IND.MEDIA VALOR',       'ferias_indenizadas_media_valor', None),
    ('GL', '4.01.01.01.30', '1/3 FERIAS RESCISAO',          'dif_terco_ferias',       None),
    ('GL', '4.01.01.01.32', '1/3 FERIAS INDENIZADAS',       'terco_ferias_indenizadas', None),
    ('GL', '4.01.01.01.33', '1/3 FERIAS PROPORCIONAIS',     'terco_ferias_prop',      None),
    ('GL', '4.01.01.01.19', '13º SALÁRIO',                  'decimo_terceiro',        None),
    ('GL', '4.01.01.01.34', '13º SAL.MEDIA VALOR',          'decimo_terceiro_media_valor', None),
    ('GL', '4.01.01.01.36', '13º SALÁRIO INDENIZADO',       'decimo_terceiro_indenizado', None),
    ('GL', '4.01.01.01.37', '13º SAL.INDEN.MEDIA VALOR',    'decimo_terceiro_indenizado_media_valor', None),
    ('GL', '4.01.01.01.38', 'AVISO PREVIO',                 'aviso_previo',           None),
    ('GL', '4.01.01.01.39', 'AVISO PREVIO MEDIA VALOR',     'aviso_previo_media_valor', None),
    ('GL', '4.01.01.01.40', 'HORAS AFAST.INSS (P/DOENC',   'horas_afastadas_inss',   None),
    ('GL', '3.01.02.02.01', 'DESC.VALE TRANSPORTE',         None, 'desc_vale_transporte'),
    ('GL', '3.01.02.02.03', 'DESC.PLANO ODONTO-SUL AMERICA',None, 'desc_plano_odonto'),
    ('GL', '3.01.02.02.04', 'DESC.PLANO SAUDE-UNIMED',      None, 'desc_plano_saude'),
    ('GL', '3.01.02.02.05', 'DESC.VA NÃO UTILIZADO',        None, None),
    ('GL', '3.01.02.02.06', 'DESC.AUX.COMBUSTIVEL 1%',      None, 'desc_aux_combustivel'),
    ('GL', '3.01.02.02.07', 'DESC.ADIANT.DE FERIAS',        None, None),
    ('GL', '4.01.01.01.40', 'HORAS AFAST.INSS (P/DOENC',   'desc_horas_afastadas',   None),
    ('GL', '3.01.02.02.08', 'DESCONTO HORAS AFASTADAS',     None, 'desc_horas_afastadas'),
    ('GL', '3.01.02.02.09', 'DESC. VALE ADIANTAMENTO',      None, 'desc_vale_adiantamento'),
    ('GL', '3.01.02.02.10', 'DESC. VALE FESTA',              None, 'desc_vale_festa'),
    ('GL', '3.01.02.02.11', 'HORAS FALTAS/ATRASOS',          None, 'desc_faltas_atrasos'),
    ('GL', '2.01.01.07.01', 'VALOR FGTS',                   None, 'valor_fgts'),
    ('GL', '2.01.01.07.06', 'DESC.INSS',                    None, 'desc_inss'),
    ('GL', '2.01.01.07.09', 'DESC.INSS FERIAS',             None, 'desc_inss_ferias'),
    ('GL', '2.01.01.07.10', 'DESC.COMPLEMENTO INSS',        None, 'desc_complemento_inss'),
    ('GL', '2.01.01.07.11', 'DESC.INSS SOBRE RESCISAO',     None, 'desc_inss_sobre_rescisao'),
    ('GL', '2.01.01.07.12', 'DESC.INSS EMPREGADOR',         None, 'desc_inss_empregador'),
    ('GL', '2.01.01.07.13', 'DESC.INSS 13 SAL.RESCISAO',    None, 'desc_inss_13sal_rescisao'),
    ('GL', '2.01.01.07.07', 'IMPOSTO DE RENDA',             None, 'desc_irrf'),
    ('GL', '4.01.01.02.03', '(-) FGTS MENSAL',               'valor_fgts', None),
]


def _lista_clts(employees):
    return [
        e for e in employees
        if e['tipo'] == 'Celetista' or e['nome'].strip().upper() in NOMES_ESPECIAIS_CLT
    ]


def _linha_e_permitida(emp, nome_c):
    if emp['nome'].strip().upper() == NOME_PROLABORE_FABIANO:
        rubricas_permitidas_fabiano = ['PRO-LABORE', 'DESC.INSS EMPREGADOR', 'LIQUIDO']
        return nome_c in rubricas_permitidas_fabiano
    if nome_c == 'PRO-LABORE':
        return False
    return True


# ════════════════════════════════════════════════════════════════════════════
# ─── EXCEL SAP (MODELO CLT) ──────────────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

def create_excel_sap(employees, data_lanc=None, data_venc=None, data_doc=None, competencia=None):
    if data_lanc is None or data_venc is None or data_doc is None or competencia is None:
        data_lanc, data_venc, data_doc, competencia = resolver_datas_sap(None)

    wb = Workbook()

    HEADER_FILL  = PatternFill('solid', fgColor='BDD7EE')
    NOME_FILL    = PatternFill('solid', fgColor='FCE4D6')
    ROW_FILL_ALT = PatternFill('solid', fgColor='EDEDED')
    ROW_FILL_WHT = PatternFill('solid', fgColor='FFFFFF')

    def hdr_font(bold=True):
        return Font(name='Calibri', bold=bold, size=10)
    def row_font(bold=False):
        return Font(name='Calibri', bold=bold, size=10)

    thin_side = Side(style='thin', color='000000')
    brd = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    ctr = Alignment(horizontal='center', vertical='center')
    lft = Alignment(horizontal='left',   vertical='center')
    rgt = Alignment(horizontal='right',  vertical='center')
    fmt_date  = 'DD/MM/YYYY'
    fmt_money = '#,##0.00'

    col_widths = [18, 12, 8, 16, 16, 16, 30, 12, 12, 20, 12, 12, 12, 10, 6, 22, 35, 14, 20, 14, 14, 40]

    clts = _lista_clts(employees)

    wb.remove(wb.active)

    for emp in clts:
        nome_aba = emp['nome'][:28].strip()
        ws = wb.create_sheet(title=nome_aba)
        ws.sheet_view.showGridLines = False

        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        headers = [
            'Campo no objeto',
            "Tipo de registro ('H' para cabeçalho e 'L' para linha)",
            'Série',
            'Data de lançamento',
            'Data de vencimento',
            'Data do documento',
            'Observações',
            'Indicador',
            'Projeto',
            'Código de transação',
            'Ref.1', 'Ref.2', 'Ref.3',
            'Tipo lançamento ECD',
            "Tipo de conta ('GL' ou 'BP')",
            'Código da conta/Código do PN',
            'Nome do PN ou conta contábil',
            'Competência da folha de pagamento',
            'Data de vencimento (em linhas)',
            'Débito',
            'Crédito',
            'Observações (nas linhas)',
        ]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font  = hdr_font()
            c.fill  = HEADER_FILL
            c.border = brd
            c.alignment = ctr
        ws.row_dimensions[1].height = 40

        ws.cell(row=2, column=2,  value='H').fill = HEADER_FILL
        ws.cell(row=2, column=4,  value=data_lanc).fill = HEADER_FILL
        ws.cell(row=2, column=5,  value=data_venc).fill = HEADER_FILL
        ws.cell(row=2, column=6,  value=data_doc).fill  = HEADER_FILL
        ws.cell(row=2, column=14, value='N').fill = HEADER_FILL
        for ci in [2, 4, 5, 6, 14]:
            c = ws.cell(row=2, column=ci)
            c.font = hdr_font(); c.border = brd; c.alignment = ctr
        for ci in [4, 5, 6]:
            ws.cell(row=2, column=ci).number_format = fmt_date
        ws.row_dimensions[2].height = 18

        codigo_bp = CODIGOS_CLT.get(emp['nome'].strip().upper(), f'F{emp["id"].zfill(6)}')
        nome_func = emp['nome']
        liquido   = emp['liquido']

        bp_vals = {
            2:  'L',
            15: 'BP',
            16: codigo_bp,
            17: nome_func,
            18: competencia,
            19: data_venc,
            22: f'{nome_func} - {competencia}',
        }
        for ci, v in bp_vals.items():
            c = ws.cell(row=3, column=ci, value=v)
            c.font   = hdr_font(bold=True)
            c.fill   = NOME_FILL
            c.border = brd
            c.alignment = lft if ci in (17, 22) else ctr
            if ci == 19:
                c.number_format = fmt_date

        ws.row_dimensions[3].height = 18

        emp_debito = 0.0
        emp_credito = 0.0

        row_num = 4
        for (tipo_c, cod_c, nome_c, chave_deb, chave_cred) in LINHAS_TEMPLATE:
            if not _linha_e_permitida(emp, nome_c):
                continue

            idx = row_num - 4
            rf = ROW_FILL_ALT if idx % 2 == 0 else ROW_FILL_WHT

            val_deb  = emp.get(chave_deb,  0.0) if chave_deb  else None
            val_cred = emp.get(chave_cred, 0.0) if chave_cred else None

            if val_deb  is not None and val_deb  == 0.0:  val_deb  = None
            if val_cred is not None and val_cred == 0.0:  val_cred = None

            gl_vals = {
                2:  'L',
                15: tipo_c,
                16: cod_c,
                17: nome_c,
                18: competencia,
                19: data_venc,
                20: val_deb,
                21: val_cred,
                22: f'{nome_func} - {competencia} - {nome_c}',
            }
            for ci, v in gl_vals.items():
                c = ws.cell(row=row_num, column=ci, value=v)
                c.font   = row_font()
                c.fill   = rf
                c.border = brd
                if ci in (20, 21) and v is not None:
                    c.alignment = rgt
                    c.number_format = fmt_money
                elif ci in (17, 22):
                    c.alignment = lft
                elif ci == 19:
                    c.alignment = ctr
                    c.number_format = fmt_date
                else:
                    c.alignment = ctr
            ws.row_dimensions[row_num].height = 18
            row_num += 1

            if val_deb is not None:
                emp_debito += val_deb
            if val_cred is not None:
                emp_credito += val_cred

        diferenca = round(abs(emp_credito - emp_debito), 2)
        if diferenca != 0:
            c = ws.cell(row=3, column=21, value=diferenca)
            c.alignment = rgt
            c.number_format = fmt_money

    if not clts:
        ws_vazio = wb.create_sheet('Sem CLTs')
        ws_vazio['A1'] = 'Nenhum funcionário CLT encontrado no PDF.'

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name)
    tmp.close()
    return tmp.name


# ════════════════════════════════════════════════════════════════════════════
# ─── INTEGRAÇÃO COM SAP BUSINESS ONE SERVICE LAYER ───────────────────────────
# ════════════════════════════════════════════════════════════════════════════

def sap_login():
    """
    Autentica no SAP Business One Service Layer.

    Segue a mesma lógica do script 2 (PJ):
      - verify=False  (certificado autoassinado)
      - Retorna os cookies da resposta (objeto RequestsCookieJar)
      - Lança exceção em caso de falha (status != 200)

    A URL é construída a partir de SAP_SL_URL (env var), igual ao script 2
    que usa a URL fixa https://b1.ativy.com:51032/b1s/v1.
    """
    url = f'{SAP_SL_URL}/Login'
    payload = {
        'CompanyDB': SAP_COMPANY_DB,
        'UserName':  SAP_USER,
        'Password':  SAP_PASSWORD,
    }
    print(f"🔑 SAP Login → {url}", flush=True)
    r = requests.post(url, json=payload, verify=False, timeout=15)
    print(f"   Status: {r.status_code}", flush=True)

    if r.status_code != 200:
        # Repassa a resposta inteira para o caller poder exibir o detalhe
        r.raise_for_status()

    # Retorna os cookies (contém B1SESSION e RouteId), igual ao script 2
    return r.cookies


def montar_lancamento_sap(employees, data_lanc=None, data_venc=None, data_doc=None, competencia=None):
    """
    Monta o payload do JournalEntry consolidando todos os CLTs.
    Datas vêm do campo escolhido no site; fallback usa DATA_LANC_PADRAO.
    """
    if data_lanc is None or data_venc is None or data_doc is None or competencia is None:
        data_lanc, data_venc, data_doc, competencia = resolver_datas_sap(None)

    clts = _lista_clts(employees)

    linhas_je = []
    funcionarios_resumo = []
    total_debito = 0.0
    total_credito = 0.0

    for emp in clts:
        codigo_bp = CODIGOS_CLT.get(emp['nome'].strip().upper(), f'F{emp["id"].zfill(6)}')
        qtd_linhas_emp = 0
        emp_debito = 0.0
        emp_credito = 0.0

        for tipo_c, cod_c, nome_c, chave_deb, chave_cred in LINHAS_TEMPLATE:
            if not _linha_e_permitida(emp, nome_c):
                continue

            val_deb  = round(emp.get(chave_deb,  0.0) or 0.0, 2) if chave_deb  else 0.0
            val_cred = round(emp.get(chave_cred, 0.0) or 0.0, 2) if chave_cred else 0.0
            if not val_deb and not val_cred:
                continue

            linhas_je.append({
                'AccountCode': cod_c,
                'ShortName':   None,
                'Debit':       val_deb,
                'Credit':      val_cred,
                'BPLID':       BPL_ID_PADRAO,
                'LineMemo':    f"{emp['nome']} - {competencia} - {nome_c}",
            })
            emp_debito    += val_deb
            emp_credito   += val_cred
            total_debito  += val_deb
            total_credito += val_cred
            qtd_linhas_emp += 1

        diferenca = round(abs(emp_credito - emp_debito), 2)
        if diferenca != 0:
            linhas_je.append({
                'AccountCode': None,
                'ShortName':   codigo_bp,
                'Debit':       0,
                'Credit':      diferenca,
                'BPLID':       BPL_ID_PADRAO,
                'LineMemo':    f"{emp['nome']} - {competencia} - DIFERENCA",
            })
            total_credito  += diferenca
            qtd_linhas_emp += 1

        funcionarios_resumo.append({
            'id':         emp['id'],
            'nome':       emp['nome'],
            'codigo_bp':  codigo_bp,
            'liquido':    diferenca,
            'qtd_linhas': qtd_linhas_emp,
        })

    # ── Formata datas no padrão SAP (YYYY-MM-DDT00:00:00), igual ao script 2 ──
    def fmt_sap(dt):
        return dt.strftime('%Y-%m-%dT00:00:00')

    payload = {
        'ReferenceDate':      fmt_sap(data_lanc),
        'DueDate':            fmt_sap(data_venc),
        'TaxDate':            fmt_sap(data_doc),
        'Memo':               f'Folha de Pagamento - {competencia}',
        'JournalEntryLines':  linhas_je,
    }
    resumo = {
        'total_funcionarios': len(funcionarios_resumo),
        'total_linhas':       len(linhas_je),
        'total_debito':       round(total_debito,  2),
        'total_credito':      round(total_credito, 2),
        'diferenca':          round(total_debito - total_credito, 2),
        'balanceado':         abs(total_debito - total_credito) < 0.01,
        'data_lancamento':    data_lanc.strftime('%d/%m/%Y'),
        'data_vencimento':    data_venc.strftime('%d/%m/%Y'),
        'data_documento':     data_doc.strftime('%d/%m/%Y'),
        'competencia':        competencia,
        'funcionarios':       funcionarios_resumo,
    }
    return payload, resumo


# ════════════════════════════════════════════════════════════════════════════
# ─── ROTAS DE CONVERSÃO (EXCEL) ──────────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════════════

@app.route('/convert', methods=['POST'])
@requer_autenticacao
def convert():
    if 'pdf' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400
    f = request.files['pdf']
    if not f.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'Envie um arquivo PDF'}), 400

    tmp_pdf = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
    f.save(tmp_pdf.name); tmp_pdf.close()

    xlsx_path = None
    try:
        employees, encargos, enc_rat = parse_pdf(tmp_pdf.name)
        if not employees:
            return jsonify({'error': 'Nenhum funcionário encontrado. Verifique se é extrato Planidata.'}), 400
        xlsx_path = create_excel(employees, encargos, enc_rat)
    finally:
        try: os.unlink(tmp_pdf.name)
        except: pass

    response = send_file(xlsx_path, as_attachment=True,
                         download_name='Folha_Pagamento.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @response.call_on_close
    def cleanup():
        try: os.unlink(xlsx_path)
        except: pass

    return response


@app.route('/convert-sap', methods=['POST'])
@requer_autenticacao
def convert_sap():
    if 'pdf' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400
    f = request.files['pdf']
    if not f.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'Envie um arquivo PDF'}), 400

    data_lanc_str = request.form.get('data_lancamento')
    data_lanc, data_venc, data_doc, competencia = resolver_datas_sap(data_lanc_str)

    tmp_pdf = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
    f.save(tmp_pdf.name); tmp_pdf.close()

    xlsx_path = None
    try:
        employees, encargos, enc_rat = parse_pdf(tmp_pdf.name)
        if not employees:
            return jsonify({'error': 'Nenhum funcionário encontrado. Verifique se é extrato Planidata.'}), 400
        xlsx_path = create_excel_sap(employees, data_lanc, data_venc, data_doc, competencia)
    finally:
        try: os.unlink(tmp_pdf.name)
        except: pass

    response = send_file(xlsx_path, as_attachment=True,
                         download_name='Folha_Pagamento_SAP.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @response.call_on_close
    def cleanup():
        try: os.unlink(xlsx_path)
        except: pass

    return response


# ════════════════════════════════════════════════════════════════════════════
# ─── ROTAS DO LANÇAMENTO NO SAP (PRÉ-VISUALIZAÇÃO + POSTAGEM) ───────────────
# ════════════════════════════════════════════════════════════════════════════

@app.route('/preview-sap', methods=['POST'])
@requer_autenticacao
def preview_sap():
    if 'pdf' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400
    f = request.files['pdf']
    if not f.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'Envie um arquivo PDF'}), 400

    data_lanc_str = request.form.get('data_lancamento')
    data_lanc, data_venc, data_doc, competencia = resolver_datas_sap(data_lanc_str)

    tmp_pdf = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
    f.save(tmp_pdf.name); tmp_pdf.close()

    try:
        employees, _, _ = parse_pdf(tmp_pdf.name)
        if not employees:
            return jsonify({'error': 'Nenhum funcionário encontrado. Verifique se é extrato Planidata.'}), 400

        payload, resumo = montar_lancamento_sap(employees, data_lanc, data_venc, data_doc, competencia)
        if resumo['total_funcionarios'] == 0:
            return jsonify({'error': 'Nenhum funcionário CLT encontrado para lançamento.'}), 400

        return jsonify({'payload': payload, 'resumo': resumo})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': f'Erro ao montar lançamento: {e}'}), 500
    finally:
        try: os.unlink(tmp_pdf.name)
        except: pass


@app.route('/postar-sap', methods=['POST'])
@requer_autenticacao
def postar_sap():
    """
    Posta o JournalEntry no SAP Business One Service Layer.

    Segue exatamente a mesma lógica do script 2 (PJ):
      1. Faz login via sap_login() → obtém cookies
      2. POST em /JournalEntries passando os cookies (não mais B1SESSION manual)
      3. verify=False para certificado autoassinado
      4. Aceita status 200 ou 201 como sucesso
    """
    body = request.get_json(silent=True) or {}
    payload = body.get('payload')
    if not payload:
        return jsonify({'error': 'Payload do lançamento não informado.'}), 400

    # Verifica balanço antes de enviar
    linhas = payload.get('JournalEntryLines', [])
    total_debito  = round(sum(float(l.get('Debit')  or 0) for l in linhas), 2)
    total_credito = round(sum(float(l.get('Credit') or 0) for l in linhas), 2)
    if abs(total_debito - total_credito) >= 0.01:
        return jsonify({
            'error': 'Lançamento não está balanceado (débito ≠ crédito). Operação cancelada.',
            'detalhe': {'total_debito': total_debito, 'total_credito': total_credito},
        }), 400

    if not (SAP_SL_URL and SAP_COMPANY_DB and SAP_USER and SAP_PASSWORD):
        return jsonify({
            'error': 'Integração com o SAP não configurada. '
                     'Defina SAP_SL_URL, SAP_COMPANY_DB, SAP_USER e SAP_PASSWORD.'
        }), 500

    # ── 1. Login no SAP (retorna cookies, igual ao script 2) ────────────────
    try:
        cookies = sap_login()
    except requests.exceptions.HTTPError as e:
        detalhe = None
        try:
            detalhe = e.response.json()
        except Exception:
            detalhe = e.response.text if e.response is not None else str(e)
        return jsonify({
            'error': 'Falha ao autenticar no SAP Service Layer '
                     '(verifique SAP_SL_URL, SAP_COMPANY_DB, SAP_USER e SAP_PASSWORD)',
            'detalhe': detalhe,
        }), 502
    except Exception as e:
        return jsonify({
            'error': 'Falha ao autenticar no SAP Service Layer',
            'detalhe': str(e),
        }), 502

    # ── 2. POST do JournalEntry (cookies passados diretamente, igual script 2) ─
    try:
        print(f"📤 Postando JournalEntry em {SAP_SL_URL}/JournalEntries", flush=True)
        resp = requests.post(
            f'{SAP_SL_URL}/JournalEntries',
            json=payload,
            cookies=cookies,          # ← cookies diretos (igual script 2)
            headers={'Content-Type': 'application/json'},
            verify=False,             # ← sem verificação SSL (igual script 2)
            timeout=30,
        )
        print(f"   Resposta SAP: {resp.status_code}", flush=True)

        if resp.status_code not in (200, 201):
            detalhe = resp.json() if resp.content else resp.text
            return jsonify({'error': 'SAP rejeitou o lançamento', 'detalhe': detalhe}), 502

        criado = resp.json()
        doc_num = criado.get('JdtNum') or criado.get('TransId')
        return jsonify({
            'mensagem': f'Lançamento criado no SAP! DocEntry: {doc_num}',
            'doc': criado,
        })

    except requests.exceptions.RequestException as e:
        return jsonify({'error': 'Erro de comunicação com o SAP', 'detalhe': str(e)}), 502


if __name__ == '__main__':
    app.run(debug=True, port=5050)